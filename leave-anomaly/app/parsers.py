"""업로드된 엑셀/CSV를 읽어 분석에 쓸 표준 형태(dict 목록)로 바꾼다.

가짜 파일을 막기 위해 확장자뿐 아니라 실제 내용(PK 시그니처·열림 여부·필수 열)까지 확인한다.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from .config import (
    ACCRUAL_COLUMNS,
    ACCRUAL_REQUIRED,
    ALLOWED_EXTENSIONS,
    USAGE_COLUMNS,
    USAGE_REQUIRED,
)

EXCEL_EPOCH = datetime(1899, 12, 30)
_DATE_RE = re.compile(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})(?:[ T]+(\d{1,2}):(\d{2}))?")


class UploadError(Exception):
    """업로드 파일이 규칙에 맞지 않을 때 발생."""


def check_extension(filename: str) -> str:
    """허용 확장자인지 확인하고 소문자 확장자를 돌려준다."""
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise UploadError(f"허용하지 않는 형식입니다. ({', '.join(sorted(ALLOWED_EXTENSIONS))} 만 가능)")
    return ext


def read_table(content: bytes, filename: str) -> list[dict]:
    """엑셀/CSV의 첫 시트를 열 이름 → 값 사전 목록으로 읽는다. 내용까지 검사한다."""
    ext = check_extension(filename)
    if ext == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]

    if not content.startswith(b"PK"):
        # .xls(구형 바이너리)는 xlsx 로 저장해 달라고 안내한다.
        raise UploadError("엑셀 파일이 아니거나 구형(.xls) 형식입니다. .xlsx 로 저장해 다시 올려주세요.")
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001 - 라이브러리 예외를 사용자 메시지로 변환
        raise UploadError("엑셀 파일을 열지 못했습니다. 파일이 손상되었는지 확인해 주세요.") from exc

    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        raise UploadError("빈 시트입니다.") from None
    names = [str(h).strip() if h is not None else "" for h in header]
    out = []
    for values in rows:
        if values is None or all(v is None or v == "" for v in values):
            continue
        out.append({names[i]: values[i] for i in range(min(len(names), len(values))) if names[i]})
    workbook.close()
    return out


def _pick(row: dict, names: list[str]) -> str:
    """후보 열 이름 중 값이 있는 첫 번째를 문자열로 돌려준다."""
    for name in names:
        value = row.get(name)
        if value is not None and str(value).strip() != "":
            return str(value).strip()
    return ""


def _raw(row: dict, names: list[str]):
    """후보 열 이름 중 값이 있는 첫 번째를 원본 타입 그대로 돌려준다."""
    for name in names:
        value = row.get(name)
        if value is not None and value != "":
            return value
    return None


def _to_number(value) -> float:
    """"1.5일" 같은 값에서 숫자만 뽑는다. 못 읽으면 0."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_cell_datetime(value) -> tuple[str, float | None]:
    """셀 값(엑셀 일련번호·datetime·문자열)을 (YYYY-MM-DD, 시각) 으로 바꾼다."""
    if value is None or value == "":
        return "", None
    if isinstance(value, datetime):
        return value.date().isoformat(), value.hour + value.minute / 60
    if isinstance(value, date):
        return value.isoformat(), None
    if isinstance(value, (int, float)):
        whole = int(value)
        dt = EXCEL_EPOCH + timedelta(days=whole)
        minutes = round((float(value) - whole) * 24 * 60)
        return dt.date().isoformat(), minutes / 60
    match = _DATE_RE.search(str(value).strip())
    if not match:
        return "", None
    y, m, d = int(match.group(1)), int(match.group(2)), int(match.group(3))
    hour = None
    if match.group(4) is not None:
        hour = int(match.group(4)) + int(match.group(5)) / 60
    return f"{y:04d}-{m:02d}-{d:02d}", hour


def _assert_columns(rows: list[dict], columns: dict, required: list[str], label: str) -> None:
    """필수 열이 실제로 있는지 확인한다(가짜 파일·다른 양식 걸러내기)."""
    if not rows:
        raise UploadError(f"{label} 파일에 데이터 행이 없습니다.")
    present = set()
    for row in rows[:50]:
        present |= {k for k, v in row.items() if v not in (None, "")}
    missing = [key for key in required if not (set(columns[key]) & present)]
    if missing:
        wanted = " / ".join(columns[key][0] for key in missing)
        raise UploadError(f"{label} 파일에서 필요한 열을 찾지 못했습니다: {wanted}")


def normalize_usage(rows: list[dict]) -> list[dict]:
    """연차 '사용내역' 시트를 분석용 표준 행으로 바꾼다. 휴가그룹이 연차휴가인 건만 남긴다."""
    _assert_columns(rows, USAGE_COLUMNS, USAGE_REQUIRED, "사용내역")
    out = []
    for row in rows:
        date_key, hour = parse_cell_datetime(_raw(row, USAGE_COLUMNS["start"]))
        item = {
            "employeeNo": _pick(row, USAGE_COLUMNS["employeeNo"]),
            "name": _pick(row, USAGE_COLUMNS["name"]),
            "org": _pick(row, USAGE_COLUMNS["org"]),
            "group": _pick(row, USAGE_COLUMNS["group"]),
            "type": _pick(row, USAGE_COLUMNS["type"]),
            "date": date_key,
            "hour": hour,
            "deducted": _to_number(_raw(row, USAGE_COLUMNS["deducted"])),
        }
        if item["name"] and item["date"] and "연차휴가" in item["group"] and item["deducted"] > 0:
            out.append(item)
    out.sort(key=lambda r: (r["name"], r["date"]))
    if not out:
        raise UploadError("휴가그룹이 '연차휴가'인 사용 기록을 찾지 못했습니다. 파일을 확인해 주세요.")
    return out


def normalize_accrual(rows: list[dict]) -> list[dict]:
    """연차 '발생내역' 시트를 분석용 표준 행으로 바꾼다."""
    _assert_columns(rows, ACCRUAL_COLUMNS, ACCRUAL_REQUIRED, "발생내역")
    out = []
    for row in rows:
        grant_start, _ = parse_cell_datetime(_raw(row, ACCRUAL_COLUMNS["grantStart"]))
        grant_end, _ = parse_cell_datetime(_raw(row, ACCRUAL_COLUMNS["grantEnd"]))
        item = {
            "employeeNo": _pick(row, ACCRUAL_COLUMNS["employeeNo"]),
            "name": _pick(row, ACCRUAL_COLUMNS["name"]),
            "org": _pick(row, ACCRUAL_COLUMNS["org"]),
            "group": _pick(row, ACCRUAL_COLUMNS["group"]),
            "state": _pick(row, ACCRUAL_COLUMNS["state"]),
            "accrued": _to_number(_raw(row, ACCRUAL_COLUMNS["accrued"])),
            "used": _to_number(_raw(row, ACCRUAL_COLUMNS["used"])),
            "remain": _to_number(_raw(row, ACCRUAL_COLUMNS["remain"])),
            "grantStart": grant_start,
            "grantEnd": grant_end,
        }
        if item["name"] and "연차휴가" in item["group"]:
            out.append(item)
    if not out:
        raise UploadError("휴가그룹이 '연차휴가'인 발생 기록을 찾지 못했습니다. 파일을 확인해 주세요.")
    return out
