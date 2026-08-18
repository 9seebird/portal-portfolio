import json, io, base64, sys
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTHS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
MONTH_KOR = ['1월', '2월', '3월', '4월', '5월', '6월', '7월', '8월', '9월', '10월', '11월', '12월']
PL = ['매출액', '매출원가', '재고자산폐기', '매출총이익', '판관비', '변동비', '광고비', '견본/판촉', '판매수수료', '운반비', '고정비', '인건비', '지급수수료', '물류수수료', '도급료', '임차료', '개발비', '포장비', '접대비', '복리후생비', '감가상각비', '기타', '공헌이익', '간접비(공통비)', '영업이익', '여비교통비', '통신비', '수도광열비', '세금과공과', '수선유지비', '보험료', '차량유지비', '소모품비', '도서인쇄비', '교육비', '행사비', '대손상각비', '사용권자산']
NPL = 38
CALC_IDX = frozenset({3, 4, 5, 22, 24, 10})
LCAT_IDX = {'매출': 0, '매출원가': 1, '재고자산폐기': 2, '광고비': 6, '견본,판촉': 7, '판매수수료': 8, '운반비': 9, '인건비': 11, '지급수수료': 12, '물류수수료': 13, '도급료': 14, '임차료': 15, '개발비': 16, '포장비': 17, '접대비': 18, '복리후생비': 19, '감가상각비': 20, '여비교통비': 25, '통신비': 26, '수도광열비': 27, '세금과공과': 28, '수선유지비': 29, '보험료': 30, '차량유지비': 31, '소모품비': 32, '도서인쇄비': 33, '교육비': 34, '행사비': 35, '대손상각비': 36, '사용권자산': 37}
KITA_N = frozenset({32, 33, 34, 35, 36, 37, 25, 26, 27, 28, 29, 30, 31})
ROW2PL = {5: 0, 6: 1, 7: 2, 8: 3, 9: 4, 10: 5, 11: 6, 12: 7, 13: 8, 14: 9, 15: 10, 16: 11, 17: 12, 18: 13, 19: 14, 20: 15, 21: 16, 22: 17, 23: 18, 24: 19, 25: 20, 26: 21, 27: 22, 28: 23, 29: 24, 34: 25, 35: 26, 36: 27, 37: 28, 38: 29, 39: 30, 40: 31, 41: 32, 42: 33, 43: 34, 44: 35, 45: 36, 46: 37}
NO_COLOR_IDX = frozenset({8})
DIRECT_IDX = {6, 7, 8, 9, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21}
KEY_ROWS = {8, 24, 3, 22}
SUB_ROWS = {10, 4, 5}
DISPLAY_ROWS = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]
DETAIL_ROWS = [25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37]
COL_H = '1F3864'
COL_HF = 'FFFFFF'
COL_SUB = 'D6E4F0'
COL_KP = 'FFF2CC'
COL_COM = 'E2EFDA'

def zero():
    return [0.0] * NPL

def safe(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.startswith("#") or not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0

def nc(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip().upper()

def ns(v):
    if v is None:
        return ""
    return str(v).strip()

def acc(pl, lcat, amt, acct_map=None, acode=None):
    if not amt:
        return
    lc = lcat
    if not lc and acode and acct_map:
        lc = acct_map.get(acode, {}).get("lcat", "")
    idx = LCAT_IDX.get(lc) if lc else None
    if idx is None:
        pl[21] += amt
    elif idx in KITA_N:
        pl[idx] += amt
        pl[21] += amt
    else:
        pl[idx] += amt

def recalc(pl):
    pl[3]  = pl[0] - pl[1] - pl[2]
    pl[5]  = sum(pl[6:10])
    pl[10] = sum(pl[11:22])
    pl[4]  = pl[5] + pl[10]
    pl[22] = pl[3] - pl[4]
    pl[24] = pl[22] - pl[23]
    return pl

def rsafe(a, b):
    return a / b if b else 0.0

def parse_alloc_targets(tstr, cg_list, alloc_map):
    """배부대상 문자열 파싱 → CG 목록 반환"""
    if not tstr:
        return []
    targets = []
    tokens  = tstr.split(",")
    for tok in tokens:
        tok    = tok.strip()
        mapped = alloc_map.get(tok, tok)   # 매핑 없으면 원본 그대로
        if mapped and mapped in cg_list and mapped not in targets:
            targets.append(mapped)
    return targets

def calc_pl(base, sdata):
    cg_list     = base["cg_list"]
    br_list     = base["br_list"]
    cc_info     = base["cc_info"]
    acct_map    = base["acct_map"]
    sales_cg    = sdata["sales_cg"]
    sales_br    = sdata["sales_br"]
    sales_cg_br = sdata["sales_cg_br"]
    cost_rate   = sdata["cost_rate"]
    exp_dir     = sdata["exp_dir"]
    exp_adv     = sdata["exp_adv"]

    pl_cg = {m: {**{cg: zero() for cg in cg_list}, "공통비": zero()} for m in MONTHS}
    pl_br = {m: {**{br: zero() for br in br_list}, "공통비": zero()} for m in MONTHS}

    # 매출액
    for m in MONTHS:
        for cg in cg_list:
            pl_cg[m][cg][0] = sales_cg[m].get(cg, 0.0)
        for br in br_list:
            pl_br[m][br][0] = sales_br[m].get(br, 0.0)

    # 매출원가
    for m in MONTHS:
        for cg in cg_list:
            for br in br_list:
                rev  = sales_cg_br[m][cg].get(br, 0.0)
                rate = cost_rate.get((cg, br, m), 0.0)
                cogs = rev * rate
                pl_cg[m][cg][1] += cogs
                pl_br[m][br][1] += cogs

    # 비용_광고판촉제외 배부
    for m in MONTHS:
        t_sales = {cg: sales_cg[m].get(cg, 0.0) for cg in cg_list}
        t_total = sum(t_sales.values())
        b_sales = {br: sales_br[m].get(br, 0.0) for br in br_list}
        b_total = sum(b_sales.values())

        for cc, pl_arr in exp_dir[m].items():
            info = cc_info.get(cc)
            if not info:
                continue
            sonik   = info["sonik"]
            alloc   = info["alloc"]
            targets = info["targets"]

            if alloc:
                tgt_cgs   = targets if targets else cg_list
                tgt_sales = {cg: t_sales.get(cg, 0.0) for cg in tgt_cgs}
                tgt_total = sum(tgt_sales.values())

                for i in range(NPL):
                    if i not in CALC_IDX:
                        pl_cg[m]["공통비"][i] += pl_arr[i]

                direct_total = sum(pl_arr[i] for i in DIRECT_IDX)
                if direct_total and tgt_total:
                    for cg in tgt_cgs:
                        r = rsafe(tgt_sales[cg], tgt_total)
                        pl_cg[m][cg][23] += direct_total * r

                for i in range(NPL):
                    if i not in CALC_IDX:
                        pl_br[m]["공통비"][i] += pl_arr[i]

                if sonik == "공통비":
                    tgt_b_sales = b_sales
                    tgt_b_total = b_total
                else:
                    tgt_b_sales = {
                        br: sum(sales_cg_br[m][cg].get(br, 0.0) for cg in tgt_cgs)
                        for br in br_list
                    }
                    tgt_b_total = sum(tgt_b_sales.values())

                if direct_total and tgt_b_total:
                    for br in br_list:
                        r = rsafe(tgt_b_sales.get(br, 0.0), tgt_b_total)
                        pl_br[m][br][23] += direct_total * r

            else:
                tgt_cg = targets[0] if targets else None
                if not tgt_cg or tgt_cg not in cg_list:
                    continue
                for i in range(NPL):
                    if i not in CALC_IDX and i != 23:
                        pl_cg[m][tgt_cg][i] += pl_arr[i]

                cb       = {br: sales_cg_br[m][tgt_cg].get(br, 0.0) for br in br_list}
                cb_total = sum(cb.values())
                for i in range(NPL):
                    if i not in CALC_IDX and i != 23:
                        v = pl_arr[i]
                        if not v:
                            continue
                        for br in br_list:
                            r = rsafe(cb[br], cb_total) if cb_total else rsafe(1, len(br_list))
                            pl_br[m][br][i] += v * r

    # 광고판촉비 배부
    for exp in exp_adv:
        ch      = exp["ch"]
        br      = exp["br"]
        lcat    = exp["lcat"]
        acode   = exp["acode"]
        monthly = exp["monthly"]

        for idx_m, amt in enumerate(monthly):
            m = idx_m + 1
            if not amt:
                continue

            if ch and br:
                if ch in cg_list:
                    acc(pl_cg[m][ch], lcat, amt, acct_map, acode)
            elif ch:
                if ch in cg_list:
                    acc(pl_cg[m][ch], lcat, amt, acct_map, acode)
            elif br:
                cg_s = {cg: sales_cg_br[m][cg].get(br, 0.0) for cg in cg_list}
                cg_t = sum(cg_s.values())
                for cg in cg_list:
                    r = rsafe(cg_s[cg], cg_t) if cg_t else rsafe(1, len(cg_list))
                    acc(pl_cg[m][cg], lcat, amt * r, acct_map, acode)
            else:
                cg_t2 = sum(sales_cg[m].values())
                for cg in cg_list:
                    r = rsafe(sales_cg[m].get(cg, 0.0), cg_t2) if cg_t2 else rsafe(1, len(cg_list))
                    acc(pl_cg[m][cg], lcat, amt * r, acct_map, acode)

            if ch and br:
                if br in br_list:
                    acc(pl_br[m][br], lcat, amt, acct_map, acode)
            elif ch:
                br_s = {b: sales_cg_br[m].get(ch, {}).get(b, 0.0) for b in br_list}
                br_t = sum(br_s.values())
                for b in br_list:
                    r = rsafe(br_s[b], br_t) if br_t else rsafe(1, len(br_list))
                    acc(pl_br[m][b], lcat, amt * r, acct_map, acode)
            elif br:
                if br in br_list:
                    acc(pl_br[m][br], lcat, amt, acct_map, acode)
            else:
                br_t2 = sum(sales_br[m].values())
                for b in br_list:
                    r = rsafe(sales_br[m].get(b, 0.0), br_t2) if br_t2 else rsafe(1, len(br_list))
                    acc(pl_br[m][b], lcat, amt * r, acct_map, acode)

    # 계산 행 업데이트
    for m in MONTHS:
        for entity in list(cg_list) + ["공통비"]:
            recalc(pl_cg[m][entity])
        for entity in list(br_list) + ["공통비"]:
            recalc(pl_br[m][entity])

    return {"pl_cg": pl_cg, "pl_br": pl_br}

def annual_data(entities, pl_by_month):
    result = {}
    for ent in entities + ["공통비"]:
        ep = zero()
        for m in MONTHS:
            src = pl_by_month[m].get(ent, zero())
            for i in range(NPL):
                if i not in CALC_IDX:
                    ep[i] += src[i]
        recalc(ep)
        result[ent] = ep
    return result

def compute_grand_total(entities, pl_dict):
    """write_cross_sheet() 의 합계 로직과 동일한 방식으로 총계 계산.
    pl_dict 는 {entity: pl배열, ..., "공통비": pl배열} 형태(월 데이터 또는 annual_data 결과)."""
    pool  = pl_dict.get("공통비", zero())
    total = zero()
    for ent in entities:
        ep = pl_dict.get(ent, zero())
        for i in range(NPL):
            if i not in CALC_IDX:
                total[i] += ep[i]
    total[23] = pool[4]   # 간접비 = 공통비풀 판관비 (BUG7과 동일 규칙)
    recalc(total)
    return total

def build_report_data(base, result):
    cg_list = base["cg_list"]
    br_list = base["br_list"]
    pl_cg   = result["pl_cg"]
    pl_br   = result["pl_br"]

    ann_cg = annual_data(cg_list, pl_cg)
    ann_br = annual_data(br_list, pl_br)

    monthly_total_cg = {m: compute_grand_total(cg_list, pl_cg[m]) for m in MONTHS}
    monthly_total_br = {m: compute_grand_total(br_list, pl_br[m]) for m in MONTHS}
    annual_total_cg  = compute_grand_total(cg_list, ann_cg)
    annual_total_br  = compute_grand_total(br_list, ann_br)

    # ── 채널 기준 합계 vs 브랜드 기준 합계 자동 검증 ─────────────
    # 매출액(0) / 공헌이익(22) / 영업이익(24) 핵심 지표만 비교
    check_idx = (0, 22, 24)
    diffs = {PL[i]: round(annual_total_cg[i] - annual_total_br[i], 2) for i in check_idx}
    ok = all(abs(v) < 1 for v in diffs.values())

    def pack(entities, pl_by_month, ann):
        return {
            "monthly": {m: {ent: pl_by_month[m].get(ent, zero()) for ent in entities + ["공통비"]}
                        for m in MONTHS},
            "annual":  {ent: ann.get(ent, zero()) for ent in entities + ["공통비"]},
        }

    return {
        "months":            MONTH_KOR,
        "pl_labels":         PL,
        "display_rows":      DISPLAY_ROWS,
        "detail_rows":       DETAIL_ROWS,
        "key_rows":          sorted(KEY_ROWS),
        "sub_rows":          sorted(SUB_ROWS),
        "no_color_idx":      sorted(NO_COLOR_IDX),
        "cg_list":           cg_list,
        "br_list":           br_list,
        "cg":                pack(cg_list, pl_cg, ann_cg),
        "br":                pack(br_list, pl_br, ann_br),
        "monthly_total_cg":  {m: monthly_total_cg[m] for m in MONTHS},
        "monthly_total_br":  {m: monthly_total_br[m] for m in MONTHS},
        "annual_total_cg":   annual_total_cg,
        "annual_total_br":   annual_total_br,
        "validation": {"diffs": diffs, "ok": ok},
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }

def hdr_fill(color):
    return PatternFill("solid", fgColor=color)

def apply_cross_style(ws, n_data_cols, last_data_row=46):
    for col in range(1, n_data_cols + 2):
        for row in (3, 4):
            c = ws.cell(row, col)
            c.fill      = hdr_fill(COL_H)
            c.font      = Font(bold=True, color=COL_HF, size=9)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_no, pl_idx in ROW2PL.items():
        lbl = ws.cell(row_no, 1)
        lbl.font = Font(bold=(pl_idx in KEY_ROWS or pl_idx in SUB_ROWS), size=9)

        if pl_idx in KEY_ROWS and pl_idx not in NO_COLOR_IDX:
            lbl.fill = hdr_fill(COL_KP)
        elif pl_idx in SUB_ROWS:
            lbl.fill = hdr_fill(COL_SUB)

        for col in range(2, n_data_cols + 2):
            c = ws.cell(row_no, col)
            if col % 2 == 1 and col > 2:
                c.number_format = "0.0%"
            else:
                c.number_format = "#,##0"
            c.font = Font(size=9)
            if pl_idx in KEY_ROWS and pl_idx not in NO_COLOR_IDX:
                c.fill = hdr_fill(COL_KP)
            elif pl_idx in SUB_ROWS:
                c.fill = hdr_fill(COL_SUB)

    ws.column_dimensions["A"].width = 16
    for col in range(2, n_data_cols + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, last_data_row + 1):
        ws.row_dimensions[row].height = 14

def apply_entity_style(ws, n_month_cols=14):
    for col in range(1, n_month_cols + 1):
        ws.cell(3, col).fill      = hdr_fill(COL_H)
        ws.cell(3, col).font      = Font(bold=True, color=COL_HF, size=9)
        ws.cell(3, col).alignment = Alignment(horizontal="center")

    for row_no, pl_idx in ROW2PL.items():
        ws.cell(row_no, 1).font = Font(bold=(pl_idx in KEY_ROWS or pl_idx in SUB_ROWS), size=9)
        for col in range(2, n_month_cols + 1):
            c = ws.cell(row_no, col)
            c.number_format = "#,##0"
            c.font          = Font(size=9)
            if pl_idx in KEY_ROWS and pl_idx not in NO_COLOR_IDX:
                c.fill = hdr_fill(COL_KP)
                ws.cell(row_no, 1).fill = hdr_fill(COL_KP)
            elif pl_idx in SUB_ROWS:
                c.fill = hdr_fill(COL_SUB)
                ws.cell(row_no, 1).fill = hdr_fill(COL_SUB)

    ws.column_dimensions["A"].width = 16
    for col in range(2, n_month_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

def write_cross_sheet(ws, title, entities, pl_data, is_annual=False):
    entity_cols = {}
    col = 2
    entity_cols["합계"] = col
    col += 2
    for ent in entities:
        entity_cols[ent] = col
        col += 2
    entity_cols["공통비"] = col

    ws.cell(1, 1).value = title
    ws.cell(3, 1).value = "항목"
    for ent, c in entity_cols.items():
        ws.cell(3, c).value = ent
        if ent != "공통비":
            ws.cell(4, c).value     = "금액"
            ws.cell(4, c + 1).value = "比"
        else:
            ws.cell(4, c).value = "금액"

    for row_no, idx in ROW2PL.items():
        ws.cell(row_no, 1).value = PL[idx]
    ws.cell(33, 1).value = ">기타 내역"
    for _r in (30, 31):
        for _c in range(1, 50):
            ws.cell(_r, _c).value = None

    # ── 합계 계산 ──────────────────────────────────────────────
    # row 6~21 (판관비 항목): CG 직접귀속분만 합산
    # row 28 (간접비):        공통비풀 판관비 전체 (pool[4])
    pool  = pl_data.get("공통비", zero())
    total = zero()
    for ent in entities:
        ep = pl_data.get(ent, zero())
        for i in range(NPL):
            if i not in CALC_IDX:
                total[i] += ep[i]

    total[23] = pool[4]   # ★ 간접비 = 공통비풀 판관비 (BUG7 수정)
    recalc(total)
    t_rev = total[0] if total[0] else 1

    c = entity_cols["합계"]
    for row_no, idx in ROW2PL.items():
        ws.cell(row_no, c).value = round(total[idx])
        ratio_cell               = ws.cell(row_no, c + 1)
        ratio_cell.value         = rsafe(total[idx], t_rev)
        ratio_cell.number_format = "0.0%"

    for ent in entities:
        c  = entity_cols[ent]
        ep = pl_data.get(ent, zero())
        e_rev = ep[0] if ep[0] else 1
        for row_no, idx in ROW2PL.items():
            ws.cell(row_no, c).value = round(ep[idx])
            ratio_cell               = ws.cell(row_no, c + 1)
            ratio_cell.value         = rsafe(ep[idx], e_rev)
            ratio_cell.number_format = "0.0%"

    c  = entity_cols["공통비"]
    cp = pl_data.get("공통비", zero())
    for row_no, idx in ROW2PL.items():
        ws.cell(row_no, c).value = round(cp[idx])

    n_cols = len(entity_cols) * 2
    apply_cross_style(ws, n_cols)

def write_entity_sheet(ws, entity_name, pl_by_month):
    ws.cell(1, 1).value = f"{entity_name}  월별 손익현황"
    ws.cell(3, 1).value = "항목"
    for m in MONTHS:
        ws.cell(3, 1 + m).value = MONTH_KOR[m - 1]
    ws.cell(3, 14).value = "연간합계"

    for row_no, idx in ROW2PL.items():
        ws.cell(row_no, 1).value = PL[idx]
    ws.cell(33, 1).value = ">기타 내역"
    for _r in (30, 31):
        for _c in range(1, 20):
            ws.cell(_r, _c).value = None

    annual = zero()
    for m in MONTHS:
        ep = pl_by_month[m].get(entity_name, zero())
        for i in range(NPL):
            if i not in CALC_IDX:
                annual[i] += ep[i]

    if entity_name == "공통비":
        annual[23] = annual[4]
    recalc(annual)

    for m in MONTHS:
        ep_orig  = pl_by_month[m].get(entity_name, zero())
        ep_write = list(ep_orig)
        if entity_name == "공통비":
            ep_write[23] = ep_orig[4]
        recalc(ep_write)
        for row_no, idx in ROW2PL.items():
            ws.cell(row_no, 1 + m).value = round(ep_write[idx])

    for row_no, idx in ROW2PL.items():
        ws.cell(row_no, 14).value = round(annual[idx])

    apply_entity_style(ws, 14)

def write_cg_file(path, base, result):
    cg_list = base["cg_list"]
    pl_cg   = result["pl_cg"]
    wb      = Workbook()
    wb.remove(wb.active)

    for m in MONTHS:
        ws = wb.create_sheet(MONTH_KOR[m - 1])
        write_cross_sheet(ws, f"{MONTH_KOR[m-1]} 고객그룹별 손익", cg_list, pl_cg[m])

    ws = wb.create_sheet("공통비")
    write_entity_sheet(ws, "공통비", pl_cg)

    ws  = wb.create_sheet("전체요약")
    ann = annual_data(cg_list, pl_cg)
    write_cross_sheet(ws, "연간합계 고객그룹별 손익", cg_list, ann, is_annual=True)

    for cg in cg_list:
        ws = wb.create_sheet(cg[:31])
        write_entity_sheet(ws, cg, pl_cg)

    wb.save(path)

def write_br_file(path, base, result):
    br_list = base["br_list"]
    pl_br   = result["pl_br"]
    wb      = Workbook()
    wb.remove(wb.active)

    for m in MONTHS:
        ws = wb.create_sheet(MONTH_KOR[m - 1])
        write_cross_sheet(ws, f"{MONTH_KOR[m-1]} 브랜드별 손익", br_list, pl_br[m])

    ws = wb.create_sheet("공통비")
    write_entity_sheet(ws, "공통비", pl_br)

    ws  = wb.create_sheet("전체요약")
    ann = annual_data(br_list, pl_br)
    write_cross_sheet(ws, "연간합계 브랜드별 손익", br_list, ann, is_annual=True)

    for br in br_list:
        ws = wb.create_sheet(br[:31])
        write_entity_sheet(ws, br, pl_br)

    wb.save(path)

def write_cc_file(path, base, sdata, result):
    cg_list = base["cg_list"]
    br_list = base["br_list"]
    cc_info = base["cc_info"]
    pl_cg   = result["pl_cg"]

    wb = Workbook()
    ws = wb.active
    ws.title = "CC배부현황"

    headers = ["CC코드", "부서", "손익분류", "배부여부", "배부대상"] + cg_list + ["공통비풀합계"]
    for j, h in enumerate(headers, 1):
        c       = ws.cell(1, j)
        c.value = h
        c.fill  = hdr_fill(COL_H)
        c.font  = Font(bold=True, color=COL_HF)

    ann_cg = annual_data(cg_list, pl_cg)

    row = 2
    for cc, info in sorted(cc_info.items()):
        common_total = sum(ann_cg["공통비"][i] for i in range(6, 22))
        vals = [
            cc, info["dept"], info["sonik"],
            "Y" if info["alloc"] else "N",
            ", ".join(info["targets"])
        ]
        for cg in cg_list:
            vals.append("")
        vals.append(round(common_total))
        for j, v in enumerate(vals, 1):
            ws.cell(row, j).value = v
        row += 1

    row += 1
    ws.cell(row, 1).value = "=== 연간 간접비(공통비) 검증 ==="
    ws.cell(row, 1).font  = Font(bold=True)
    row += 1
    ws.cell(row, 1).value = "구분"
    ws.cell(row, 2).value = "간접비(공통비)"
    ws.cell(row, 1).font  = Font(bold=True)
    ws.cell(row, 2).font  = Font(bold=True)
    row += 1

    total_indirect = 0
    for cg in cg_list:
        ws.cell(row, 1).value = cg
        v = round(ann_cg[cg][23])
        ws.cell(row, 2).value = v
        total_indirect += v
        row += 1

    ws.cell(row, 1).value = "CG 간접비 합계"
    ws.cell(row, 2).value = total_indirect
    ws.cell(row, 1).font  = Font(bold=True)
    row += 1

    common_pool = round(ann_cg["공통비"][4])
    ws.cell(row, 1).value = "공통비 풀 판관비"
    ws.cell(row, 2).value = common_pool
    ws.cell(row, 1).font  = Font(bold=True)
    row += 1

    diff = total_indirect - common_pool
    ws.cell(row, 1).value = "검증 (차이)"
    ws.cell(row, 2).value = "일치 ✅" if diff == 0 else f"차이: {diff:,}"
    ws.cell(row, 1).font  = Font(bold=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    wb.save(path)

def build_raw_payload(base, sdata):
    cg_list = base["cg_list"]
    br_list = base["br_list"]
    sales_cg_br = sdata["sales_cg_br"]
    cost_rate   = sdata["cost_rate"]
    exp_dir     = sdata["exp_dir"]
    exp_adv     = sdata["exp_adv"]

    cost_rate_nested = {cg: {br: {} for br in br_list} for cg in cg_list}
    for (cg, br, m), rate in cost_rate.items():
        if cg in cost_rate_nested and br in cost_rate_nested[cg]:
            cost_rate_nested[cg][br][str(m)] = rate

    exp_dir_j = {}
    for m in MONTHS:
        exp_dir_j[str(m)] = {}
        for cc, arr in exp_dir[m].items():
            arr2 = list(arr)
            kita_sum   = sum(arr2[i] for i in KITA_N)
            arr2[21]   = arr2[21] - kita_sum
            exp_dir_j[str(m)][cc] = arr2

    sales_cg_br_j = {
        str(m): {cg: dict(sales_cg_br[m][cg]) for cg in cg_list}
        for m in MONTHS
    }

    return {
        "cg_list":  cg_list,
        "br_list":  br_list,
        "cc_info":  {cc: dict(info) for cc, info in base["cc_info"].items()},
        "sales_cg_br": sales_cg_br_j,
        "cost_rate":   cost_rate_nested,
        "exp_dir":     exp_dir_j,
        "exp_adv":     exp_adv,
    }

def base_from_payload(payload):
    return {
        "cg_list":  payload["cg_list"],
        "br_list":  payload["br_list"],
        "cc_info":  payload["cc_info"],
        "acct_map": {},
    }

def payload_to_runtime(payload):
    cg_list = payload["cg_list"]
    br_list = payload["br_list"]

    sales_cg_br = {
        m: {
            cg: {br: float(payload["sales_cg_br"][str(m)][cg].get(br, 0.0)) for br in br_list}
            for cg in cg_list
        }
        for m in MONTHS
    }
    sales_cg = {m: {cg: sum(sales_cg_br[m][cg].values()) for cg in cg_list} for m in MONTHS}
    sales_br = {m: {br: sum(sales_cg_br[m][cg][br] for cg in cg_list) for br in br_list} for m in MONTHS}

    cost_rate = {}
    for cg in cg_list:
        for br in br_list:
            for m in MONTHS:
                v = payload["cost_rate"].get(cg, {}).get(br, {}).get(str(m))
                if v is not None:
                    cost_rate[(cg, br, m)] = float(v)

    exp_dir = {}
    for m in MONTHS:
        exp_dir[m] = {}
        for cc, arr in payload["exp_dir"][str(m)].items():
            arr2 = [float(x) for x in arr]
            kita_sum = sum(arr2[i] for i in KITA_N)
            arr2[21] = arr2[21] + kita_sum
            exp_dir[m][cc] = arr2

    exp_adv = payload["exp_adv"]

    return {
        "sales_cg": sales_cg, "sales_br": sales_br, "sales_cg_br": sales_cg_br,
        "cost_rate": cost_rate, "exp_dir": exp_dir, "exp_adv": exp_adv,
    }

def recompute(payload_json):
    payload = json.loads(payload_json)
    base    = base_from_payload(payload)
    sdata   = payload_to_runtime(payload)
    result  = calc_pl(base, sdata)
    data    = build_report_data(base, result)
    return json.dumps(data, ensure_ascii=False)

def export_files(payload_json):
    payload = json.loads(payload_json)
    base    = base_from_payload(payload)
    sdata   = payload_to_runtime(payload)
    result  = calc_pl(base, sdata)

    out = []
    for fn in (write_cg_file, write_br_file):
        buf = io.BytesIO()
        fn(buf, base, result)
        out.append(base64.b64encode(buf.getvalue()).decode())

    buf3 = io.BytesIO()
    write_cc_file(buf3, base, sdata, result)
    out.append(base64.b64encode(buf3.getvalue()).decode())

    return json.dumps(out)


# ── 채널 이름 별칭 ───────────────────────────────────────────
# 두 파일이 같은 채널을 다르게 부르는 일이 흔하다 (예: 마스터는 「헬스뷰티」,
# 매출목표는 「H&B」). 그 짝을 **마스터 파일 「고객그룹」 시트 C열**에 적는다.
#
# 코드에 미리 박아 두지 않는 이유 —
#   · 채널이 늘거나 이름이 바뀔 때마다 코드를 고치고 다시 배포해야 한다
#   · 무엇보다, 코드에 적힌 이름과 파일에 적힌 이름이 어긋나면
#     그 채널 매출이 **오류 없이 통째로 빠진다**
# 파일 한 곳에서만 관리하면 이 두 가지가 다 없어진다.
#
# 여기 채워 넣으면 파일에 안 적혀 있어도 알아듣는다. 기본은 비워 둔다.
CHANNEL_MAP_FALLBACK = {}
CC_SONIK_FALLBACK = {}

def load_base(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── 고객그룹 시트 ──────────────────────────────────────────
    # A열: 코드, B열: CG명, C열(선택): 채널 별칭(쉼표 구분)
    ws = wb["고객그룹"]
    cg_list     = []
    channel_map = dict(CHANNEL_MAP_FALLBACK)   # hardcoded fallback 먼저 넣음
    alloc_map   = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cg_name = ns(row[1]) if len(row) > 1 else ""
        if not cg_name:
            continue
        cg_list.append(cg_name)

        # CG명 자체를 self-mapping (★ 신규 CG 자동 인식 핵심)
        channel_map[cg_name] = cg_name
        alloc_map[cg_name]   = cg_name

        # C열: 추가 채널 별칭 (기입돼 있을 때만)
        if len(row) > 2 and row[2]:
            aliases = str(row[2]).split(",")
            for alias in aliases:
                alias = alias.strip()
                if alias:
                    channel_map[alias] = cg_name
                    alloc_map[alias]   = cg_name

    # alloc_map 에도 fallback 병합 (단, 파일 쪽 우선)
    for k, v in CHANNEL_MAP_FALLBACK.items():
        if k not in alloc_map:
            alloc_map[k] = v

    # ── 브랜드 시트 ───────────────────────────────────────────
    ws = wb["브랜드"]
    br_list = [ns(r[0]) for i, r in enumerate(ws.iter_rows(values_only=True))
               if i > 0 and ns(r[0])]

    # ── 계정명 시트 ───────────────────────────────────────────
    ws = wb["계정명"]
    acct_map = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        code = ns(row[0])
        if code:
            acct_map[code] = {"name": ns(row[1]), "lcat": ns(row[2])}

    # ── 코스트센터 시트 ────────────────────────────────────────
    # A열: CC코드, B열: 부서, C열: 손익구분, D열: 배부여부, E열: 배부대상
    ws = wb["코스트센터"]
    cc_info = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cc = nc(row[0])
        if not cc:
            continue
        sonik = ns(row[2])
        alloc = ns(row[3]).upper() == "Y"
        tstr  = ns(row[4]) if len(row) > 4 and row[4] else ""

        if alloc:
            targets = parse_alloc_targets(tstr, cg_list, alloc_map)
            if not targets:
                targets = cg_list[:]
        else:
            # ★ 손익구분이 cg_list에 직접 있으면 바로 사용 (파일 수정만으로 해결)
            if sonik in cg_list:
                cg = sonik
            else:
                # fallback: CC_SONIK_FALLBACK
                cg = CC_SONIK_FALLBACK.get(sonik, "")
                if not cg:
                    # channel_map에서도 찾기
                    cg = channel_map.get(sonik, "")
            targets = [cg] if cg and cg in cg_list else []

        cc_info[cc] = {
            "dept":    ns(row[1]),
            "sonik":   sonik,
            "alloc":   alloc,
            "targets": targets,
        }

    return {
        "cg_list":     cg_list,
        "br_list":     br_list,
        "acct_map":    acct_map,
        "cc_info":     cc_info,
        "channel_map": channel_map,   # ★ 동적 생성
        "alloc_map":   alloc_map,     # ★ 동적 생성
    }

def load_sales(path, cg_list, br_list, acct_map, channel_map):
    wb = openpyxl.load_workbook(path, data_only=True)

    ws   = wb["매출목표"]
    rows = list(ws.iter_rows(values_only=True))

    sales_cg    = {m: {cg: 0.0 for cg in cg_list} for m in MONTHS}
    sales_br    = {m: {br: 0.0 for br in br_list}  for m in MONTHS}
    sales_cg_br = {m: {cg: {br: 0.0 for br in br_list} for cg in cg_list} for m in MONTHS}

    in_sec4 = False
    cur_ch  = None

    for row in rows:
        v1 = ns(row[1]) if len(row) > 1 and row[1] else ""

        if "4)" in v1 and "채널" in v1:
            in_sec4 = True
            continue
        if not in_sec4:
            continue

        col1 = ns(row[1]) if len(row) > 1 and row[1] else ""
        col2 = ns(row[2]) if len(row) > 2 and row[2] else ""

        if col1 in ("합 계", "합계") and col2 in ("", "None"):
            break

        # ── 채널명 갱신 ──────────────────────────────────────
        # ★ BUG8 수정: 채널명 행이지만 매핑 실패 시 cur_ch = None 으로 리셋
        #   (이전 CG 유지로 인한 오귀속 방지)
        if col1 and col1 not in ("합 계", "합계", "브랜드", "채널", "None", ""):
            mapped = channel_map.get(col1)
            if mapped and mapped in cg_list:
                cur_ch = mapped
            else:
                # 인식 못한 채널명 → 다음 인식 가능한 채널 전까지 skip
                cur_ch = None

        if cur_ch is None:
            continue

        if col2 in ("합계", "합 계"):
            for m in MONTHS:
                v = safe(row[4 + m])
                sales_cg[m][cur_ch] += v
        elif col2 and col2 not in ("比", "채널", "브랜드", "비고", "None", ""):
            br_name = col2 if col2 in br_list else None
            if br_name:
                for m in MONTHS:
                    v = safe(row[4 + m])
                    sales_cg_br[m][cur_ch][br_name] += v
                    sales_br[m][br_name]             += v

    # ── 매출원가율 ────────────────────────────────────────────
    ws        = wb["매출원가율"]
    cost_rate = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        col1 = ns(row[1]) if row[1] else ""
        col2 = ns(row[2]) if row[2] else ""
        if not col1 or col1 in ("합 계", "합계", "브랜드"):
            continue
        if col2 in ("합계", "합 계", ""):
            continue
        ch_cg = channel_map.get(col1)
        if not ch_cg or ch_cg not in cg_list:
            continue
        br_name = col2 if col2 in br_list else None
        if not br_name:
            continue
        for m in MONTHS:
            rate = safe(row[2 + m])
            cost_rate[(ch_cg, br_name, m)] = rate

    # ── 비용_광고판촉제외 ──────────────────────────────────────
    ws      = wb["비용_광고판촉제외"]
    exp_dir = {m: {} for m in MONTHS}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cc = nc(row[0])
        if not cc:
            continue
        acode    = nc(row[1])
        lcat_raw = row[16] if len(row) > 16 else None
        if lcat_raw is None or str(lcat_raw).startswith("#"):
            lcat = ""
        else:
            lcat = ns(lcat_raw)
        if not lcat and acode:
            lcat = acct_map.get(acode, {}).get("lcat", "")

        for m in MONTHS:
            v = safe(row[3 + m])
            if not v:
                continue
            if cc not in exp_dir[m]:
                exp_dir[m][cc] = zero()
            acc(exp_dir[m][cc], lcat, v, acct_map, acode)

    # ── 비용_광고판촉,개발 ────────────────────────────────────
    ws      = wb["비용_광고판촉,개발"]
    exp_adv = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cc = nc(row[0])
        if not cc:
            continue
        ch_raw    = ns(row[1]) if row[1] else ""
        br_raw    = ns(row[2]) if row[2] else ""
        acode     = nc(row[3]) if row[3] else ""
        lcat_raw2 = row[18] if len(row) > 18 else None
        if lcat_raw2 is None or str(lcat_raw2).startswith("#"):
            lcat = ""
        else:
            lcat = ns(lcat_raw2)
        if not lcat and acode:
            lcat = acct_map.get(acode, {}).get("lcat", "")

        monthly = [safe(row[5 + m]) for m in MONTHS]

        ch = channel_map.get(ch_raw) if ch_raw not in ("공통", "") else None
        br = br_raw if (br_raw and br_raw not in ("공통", "") and br_raw in br_list) else None

        exp_adv.append({
            "cc": cc, "ch": ch, "br": br,
            "lcat": lcat, "acode": acode, "monthly": monthly,
        })

    return {
        "sales_cg":  sales_cg,  "sales_br":  sales_br,
        "sales_cg_br": sales_cg_br,
        "cost_rate": cost_rate, "exp_dir":   exp_dir,
        "exp_adv":   exp_adv,
    }


def init_report_from_b64(base_b64, sales_b64):
    base_bytes = base64.b64decode(base_b64)
    sales_bytes = base64.b64decode(sales_b64)
    base = load_base(io.BytesIO(base_bytes))
    sdata = load_sales(io.BytesIO(sales_bytes), base["cg_list"], base["br_list"], base["acct_map"], base["channel_map"])
    result = calc_pl(base, sdata)
    data = build_report_data(base, result)
    raw_pl = build_raw_payload(base, sdata)
    return json.dumps({"data": data, "raw_payload": raw_pl}, ensure_ascii=False)
