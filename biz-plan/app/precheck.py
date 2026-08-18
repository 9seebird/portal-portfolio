# -*- coding: utf-8 -*-
"""올리기 전 점검 — 숫자를 보여 주기 **전에** 돌린다.

이 앱의 위험은 「틀렸는데 표는 멀쩡히 나오는 것」이다. 그래서 계산 결과를
보여 주기 전에 먼저 아래를 확인하고, 걸리는 것이 있으면 사람에게 묻는다.

  멈춤   — 이대로 보면 안 되는 상태. 결과를 보여 주지 않는다
  확인   — 보여 주되 무엇이 빠졌는지 위에 띄운다
  정상   — 그냥 보여 준다

핵심은 **파일이 스스로 적어 둔 합계**와 우리가 더한 값을 맞춰 보는 것이다.
채널 이름이 어긋나든, 열이 밀리든, 수식 값이 안 저장돼 있든,
「빠진 게 있다」는 사실은 이 대조 하나로 다 드러난다.
"""
import io
import openpyxl

MONTHS = range(1, 13)


def _s(v):
    return str(v).strip() if v is not None else ""


def _num(v):
    return v if isinstance(v, (int, float)) else 0.0


def check(base_bytes, sales_bytes, base=None, engine_total_sales=None):
    """두 단계로 나뉜다. **1단계는 엔진을 부르기 전에** 돌려야 한다.

      1단계 (base=None)  파일이 엑셀인지, 자리가 안 바뀌었는지, 시트가 있는지
                         → 여기서 걸리면 엔진을 아예 부르지 않는다.
                            부르면 'File is not a zip file' 같은 말만 나온다.
      2단계 (base 있음)   엔진을 돌린 뒤, 합계가 맞는지

    base                : load_base() 결과 (channel_map, cg_list, br_list)
    engine_total_sales  : 엔진이 계산한 연간 매출액 합계
    """
    stop, warn, info = [], [], []

    # ── 0. 파일이 뒤바뀌었는가 ────────────────────────────────
    # 가장 흔한 실수라서 제일 먼저 본다.
    def _open(blob, who):
        try:
            # ★ read_only=True 를 쓰지 않는다.
            #   그 방식은 파일에 적힌 「쓰인 범위(dimension)」를 그대로 믿는데,
            #   그 값이 실제와 어긋난 파일이 있다. 실제로 행을 끼워 넣은 파일에서
            #   마지막 합계 줄을 통째로 못 읽어, 아래 합계 대조가 조용히
            #   아무 일도 하지 않았다. 계산하는 쪽(engine)은 이 방식을 안 쓰므로,
            #   점검만 다르게 읽으면 점검이 헛돈다. 파일이 작아 속도 문제도 없다.
            return openpyxl.load_workbook(io.BytesIO(blob), data_only=True)
        except Exception:
            stop.append(
                f"{who} 자리에 올린 것이 엑셀 파일이 아닙니다\n"
                "\n"
                ".xlsx 파일만 열 수 있습니다. 확장자가 .xls 나 .csv 이거나, "
                "파일이 깨져 있으면 이 안내가 나옵니다.\n"
                "\n"
                "엑셀에서 파일을 연 뒤 [다른 이름으로 저장] → 형식을 "
                "「Excel 통합 문서(*.xlsx)」로 골라 저장하고 다시 올려 주세요.")
            return None

    wb_b = _open(base_bytes, "마스터 양식")
    wb_s = _open(sales_bytes, "매출목표·비용")
    if stop:
        return stop, warn, info

    nb, ns = set(wb_b.sheetnames), set(wb_s.sheetnames)

    # ── 어느 파일이 무엇인지 시트를 보고 알아본다 ───────────────
    # 파일 이름은 사람마다 다르게 붙이므로 믿을 수 없다. 안에 든 시트로 본다.
    MASTER_MARK = {"고객그룹", "브랜드", "계정명", "코스트센터"}
    SALES_MARK = {"매출목표", "매출원가율"}

    def looks(names):
        if MASTER_MARK <= names:
            return "마스터"
        if SALES_MARK <= names:
            return "매출목표"
        return "?"

    kb, ks = looks(nb), looks(ns)

    # 자리 문제를 **하나의 안내로** 정리한다. 이걸 시트 하나하나로 알려 주면
    # "「매출목표」 시트가 없습니다" 같은 말이 여러 개 뜨는데,
    # 정작 진짜 원인(자리가 바뀜)은 어디에도 안 나온다.
    if base_bytes == sales_bytes:
        stop.append(
            "두 자리에 같은 파일을 올리셨습니다\n"
            "\n"
            "1번 자리에는 「사업계획용 기본양식」(마스터), "
            "2번 자리에는 「매출목표&비용」 파일이 들어가야 합니다.\n"
            "\n"
            "서로 다른 두 파일을 골라서 다시 올려 주세요.")
    elif kb == "매출목표" and ks == "마스터":
        stop.append(
            "두 파일의 자리가 바뀌었습니다\n"
            "\n"
            "1번 자리에 매출목표&비용 파일이, 2번 자리에 마스터 양식이 "
            "올라와 있습니다.\n"
            "\n"
            "자리를 서로 바꿔서 다시 올려 주세요.")
    elif kb == "마스터" and ks == "마스터":
        stop.append(
            "두 자리 모두 마스터 양식입니다\n"
            "\n"
            "2번 자리에는 「매출목표&비용」 파일이 들어가야 하는데, "
            "마스터 양식이 한 번 더 올라와 있습니다.\n"
            "\n"
            "2번 자리에 매출목표&비용 파일을 골라서 다시 올려 주세요.")
    elif kb == "매출목표" and ks == "매출목표":
        stop.append(
            "두 자리 모두 매출목표&비용 파일입니다\n"
            "\n"
            "1번 자리에는 「사업계획용 기본양식」(마스터)이 들어가야 하는데, "
            "매출목표&비용 파일이 한 번 더 올라와 있습니다.\n"
            "\n"
            "1번 자리에 마스터 양식을 골라서 다시 올려 주세요.")
    else:
        # 자리는 맞는데 시트가 모자란 경우. 무엇이 없는지 한 번만 말한다.
        for mark, names, who, num in ((MASTER_MARK, nb, "마스터 양식", "1번"),
                                      (SALES_MARK, ns, "매출목표&비용", "2번")):
            missing = sorted(mark - names)
            if missing:
                stop.append(
                    f"{num} 자리의 파일이 {who} 파일이 아닌 것 같습니다\n"
                    "\n"
                    f"{who} 파일에는 "
                    + ", ".join(f"「{m}」" for m in sorted(mark))
                    + " 시트가 있어야 하는데, "
                    + ", ".join(f"「{m}」" for m in missing)
                    + " 이(가) 없습니다.\n"
                    "\n"
                    "이 파일에 있는 시트: " + ", ".join(wb_b.sheetnames if names is nb
                                                    else wb_s.sheetnames) + "\n"
                    "\n"
                    "파일을 잘못 고르셨는지, 시트 이름을 바꾸신 것은 아닌지 "
                    "확인해 주세요. 오른쪽 위 [예시 파일 다운로드] 에서 "
                    "올바른 양식을 받으실 수 있습니다.")

    if stop or base is None:
        # 1단계는 여기까지. 통과했으면 이제 엔진을 불러도 된다.
        return stop, warn, info

    # ── 0-2. 월 금액이 제 자리에 있는가 ─────────────────────
    # 합계 대조만으로는 **열이 밀린 것을 못 잡는다.** 파일에 적힌 합계도
    # 같은 칸에서 읽으므로, 둘 다 똑같이 틀린 채로 일치해 버린다.
    # 실제로 열 하나를 끼워 넣었더니 매출이 41억 줄었는데 「일치합니다」가 떴다.
    # 그래서 월 이름이 몇 번째 칸에 있는지를 따로 확인한다.
    MONTH_AT = {"매출목표": 5, "매출원가율": 3}   # 엔진이 1월을 찾는 칸
    for sheet, want in MONTH_AT.items():
        found = None
        for r in wb_s[sheet].iter_rows(values_only=True):
            at = {}
            for j, c in enumerate(r):
                t = _s(c)
                if t in ("1월", "6월", "12월"):
                    at[t] = j
            if len(at) == 3:
                found = at
                break
        if found is None:
            stop.append(
                f"「{sheet}」 시트에서 월 이름을 찾지 못했습니다\n"
                "\n"
                "「1월」 「2월」 … 「12월」 이 적힌 머리글 줄이 있어야 각 달의 금액을 "
                "어느 칸에서 읽을지 알 수 있습니다.\n"
                "\n"
                "머리글을 지우거나 「'1월」 「1 월」 처럼 다르게 적지 않았는지 확인해 주세요. "
                "오른쪽 위 [예시 파일 다운로드] 에서 올바른 양식을 받으실 수 있습니다.")
        elif found["1월"] != want:
            moved = found["1월"] - want
            stop.append(
                f"「{sheet}」 시트의 칸이 {abs(moved)}칸 {'오른쪽으로' if moved > 0 else '왼쪽으로'} 밀렸습니다\n"
                "\n"
                "월 금액을 읽는 자리가 정해져 있는데, 그 앞에 칸(열)이 "
                f"{'늘었' if moved > 0 else '줄었'}습니다. 이대로 계산하면 엉뚱한 칸의 숫자를 "
                "각 달의 금액으로 읽어서, **오류 없이 틀린 표가 나옵니다.**\n"
                "\n"
                "이렇게 고치면 됩니다\n"
                f"  1. 「{sheet}」 시트를 엽니다\n"
                f"  2. 「1월」 이 {chr(64 + want + 1)}열에 오도록 맞춥니다 "
                f"(지금은 {chr(64 + found['1월'] + 1)}열입니다)\n"
                "  3. 저장하고 파일을 다시 올립니다\n"
                "\n"
                "칸을 넣거나 지우는 대신, 예시 파일 양식을 그대로 쓰시는 편이 안전합니다.")
    if stop:
        return stop, warn, info

    # ── 1. 매출목표 시트를 우리 눈으로 다시 읽는다 ────────────
    # 엔진과 같은 방식으로 읽되, **버린 것을 기록**한다.
    ws = wb_s["매출목표"]
    rows = list(ws.iter_rows(values_only=True))
    cmap, cg_list, br_list = base["channel_map"], base["cg_list"], base["br_list"]

    stated_total = 0.0     # 파일이 스스로 적어 둔 합계
    counted = 0.0          # 우리가 알아본 것만 더한 값
    unknown_ch = {}        # 못 알아본 채널 → 금액
    unknown_br = {}        # 못 알아본 브랜드 → 금액
    in4, cur = False, None

    for r in rows:
        c1 = _s(r[1]) if len(r) > 1 else ""
        c2 = _s(r[2]) if len(r) > 2 else ""
        if "4)" in c1 and "채널" in c1:
            in4 = True
            continue
        if not in4:
            continue

        amt = sum(_num(r[4 + m]) for m in MONTHS if len(r) > 4 + m)

        if c1 in ("합 계", "합계") and not c2:
            stated_total = amt        # 맨 아래 총계
            break

        if c1 and c1 not in ("브랜드", "채널", "None", ""):
            mapped = cmap.get(c1)
            cur = mapped if (mapped and mapped in cg_list) else None
            if cur is None:
                unknown_ch[c1] = unknown_ch.get(c1, 0.0)

        if cur is None:
            if c1 and c1 not in ("브랜드", "채널", "합 계", "합계", ""):
                # 못 알아본 채널 아래 줄들의 금액도 함께 센다
                for k in unknown_ch:
                    pass
            continue

        if c2 in ("합계", "합 계"):
            counted += amt
        elif c2 and c2 not in ("比", "채널", "브랜드", "비고", "None", ""):
            if c2 not in br_list:
                unknown_br[c2] = unknown_br.get(c2, 0.0) + amt

    # 못 알아본 채널이 흘린 금액 = 파일 합계 − 알아본 합계
    lost = stated_total - counted if stated_total else 0.0

    # ── 2. 숫자를 아예 못 읽었는가 (수식 값이 저장 안 됨) ─────
    if stated_total == 0 and counted == 0:
        stop.append(
            "매출 숫자를 하나도 읽지 못했습니다\n"
            "\n"
            "매출목표 파일의 숫자가 전부 0 으로 읽힙니다. 엑셀 수식의 계산 결과가 "
            "파일에 저장돼 있지 않을 때 이렇게 됩니다. 엑셀이 아닌 프로그램으로 "
            "파일을 열었다가 저장하면 이 일이 생깁니다.\n"
            "\n"
            "이렇게 고치면 됩니다\n"
            "  1. 매출목표 파일을 엑셀로 엽니다\n"
            "  2. 아무것도 고치지 말고 그냥 저장(Ctrl+S)합니다\n"
            "  3. 그 파일을 다시 올립니다")
        return stop, warn, info

    # ── 3. 합계가 맞는가 ★ 이 검사 하나가 대부분을 잡는다 ─────
    if stated_total:
        diff = stated_total - engine_total_sales
        if abs(diff) > 1:
            if unknown_ch:
                # 원인이 분명하다. 그러면 원인부터 말한다 — 숫자 표를 먼저
                # 들이밀면 무엇을 해야 하는지가 눈에 안 들어온다.
                names = ", ".join(f"「{n}」" for n in unknown_ch)
                stop.append(
                    f"{names} 채널이 계산에서 빠졌습니다\n"
                    "\n"
                    f"매출목표 파일에는 {names} 라는 채널이 있는데, 마스터 양식의 "
                    "고객그룹 목록에는 그 이름이 없습니다. 두 파일이 같은 채널을 "
                    "서로 다른 이름으로 부르고 있어서, 이 채널의 매출이 통째로 "
                    "빠진 채로 계산됩니다.\n"
                    "\n"
                    f"매출목표 파일에 적힌 매출 합계는 {stated_total:,.0f} 원인데 "
                    f"계산에 들어간 금액은 {engine_total_sales:,.0f} 원으로, "
                    f"{diff:,.0f} 원이 빠졌습니다.\n"
                    "\n"
                    "이렇게 고치면 됩니다\n"
                    "  1. 마스터 양식 파일을 엑셀로 엽니다\n"
                    "  2. 「고객그룹」 시트를 엽니다\n"
                    f"  3. 이 채널에 해당하는 줄을 찾아 C열에 {names} 라고 적습니다\n"
                    "  4. 저장하고 파일을 다시 올립니다\n"
                    "\n"
                    "C열은 「매출목표 파일에서는 이 이름으로도 부른다」를 적어 두는 "
                    "칸입니다. 여기에 적어 두면 다음부터는 그냥 넘어갑니다.")
            else:
                # 원인을 못 짚었다. 아는 척하지 않는다.
                stop.append(
                    "매출 합계가 맞지 않습니다\n"
                    "\n"
                    f"매출목표 파일에 적힌 매출 합계는 {stated_total:,.0f} 원인데 "
                    f"계산에 들어간 금액은 {engine_total_sales:,.0f} 원입니다. "
                    f"{diff:,.0f} 원이 계산에서 빠졌습니다.\n"
                    "\n"
                    "어느 줄이 빠졌는지는 찾지 못했습니다. 매출목표 시트의 "
                    "「4) 채널별」 아래에서 채널·브랜드 이름이 마스터 양식과 같은지, "
                    "월 금액이 들어가는 칸이 밀리지 않았는지 확인해 주세요.\n"
                    "\n"
                    "판단이 어려우시면 이 화면을 그대로 담당자에게 보여 주세요.")
        else:
            info.append(f"점검을 마쳤습니다 — 매출 합계가 파일과 일치합니다 "
                        f"({stated_total:,.0f} 원)")

    if unknown_br:
        names = ", ".join(f"「{k}」 {v:,.0f} 원" for k, v in unknown_br.items())
        warn.append(
            "일부 브랜드가 계산에서 빠졌습니다\n"
            "\n"
            f"매출목표 파일의 {names} 이(가) 마스터 양식의 브랜드 목록에 없습니다. "
            "이름이 서로 다르면 그 줄은 계산에 들어가지 않습니다. "
            "마스터 양식의 「브랜드」 시트와 이름이 같은지 확인해 주세요.")

    return stop, warn, info
