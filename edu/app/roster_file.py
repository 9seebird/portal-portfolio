"""명단 파일(CSV·엑셀) 읽기.

사람마다 열 이름을 다르게 적는다. '사번'일 수도 '사원번호'일 수도 있어서
config.ROSTER_COLUMNS 의 후보 목록으로 찾는다. 잘못된 파일은 **왜** 잘못됐는지
알려 준다. "형식이 올바르지 않습니다"만 나오면 고칠 수가 없다.
"""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from .config import ROSTER_COLUMNS, ROSTER_EXTENSIONS, ROSTER_MAX_BYTES

TRUE_WORDS = {"1", "y", "yes", "true", "o", "ox", "예", "네", "담당", "담당자", "관리자", "admin", "v", "✓"}


class RosterFileError(ValueError):
    """명단 파일을 읽지 못했을 때. 사람이 읽고 고칠 수 있는 이유를 담는다."""


def _norm_header(s: str) -> str:
    """열 이름 비교용. 공백·괄호·기호를 털어내고 소문자로."""
    return re.sub(r"[\s()\[\]/_·\-]", "", str(s or "")).lower()


def _find_columns(header: list[str]) -> dict[str, int]:
    """첫 줄에서 사번·이름·부서·담당자 열이 몇 번째인지 찾는다."""
    norm = [_norm_header(h) for h in header]
    found: dict[str, int] = {}
    for key, candidates in ROSTER_COLUMNS.items():
        for cand in candidates:
            c = _norm_header(cand)
            if c in norm:
                found[key] = norm.index(c)
                break
    return found


def _read_csv(content: bytes) -> list[list[str]]:
    """CSV 를 읽는다. 엑셀에서 저장한 한글 CSV(cp949)도 받아 준다."""
    text = None
    for enc in ("utf-8-sig", "cp949", "utf-8"):
        try:
            text = content.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise RosterFileError("글자 인코딩을 알 수 없습니다. 엑셀에서 'CSV UTF-8'로 다시 저장해 주세요.")
    try:
        dialect = csv.Sniffer().sniff(text[:2000], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect) if any(str(c).strip() for c in row)]


def _read_xlsx(content: bytes) -> list[list[str]]:
    """엑셀 파일의 첫 시트를 읽는다."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - requirements 에 들어 있다
        raise RosterFileError("엑셀을 읽을 수 없는 환경입니다. CSV 로 올려 주세요.") from exc
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise RosterFileError("엑셀 파일을 열지 못했습니다. 파일이 손상되었을 수 있습니다.") from exc
    ws = wb.worksheets[0]
    rows = []
    for r in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in r]
        if any(cells):
            rows.append(cells)
    wb.close()
    return rows


def parse(filename: str, content: bytes) -> list[dict]:
    """명단 파일을 읽어 [{id, name, dept, is_admin, active}, ...] 로 돌려준다."""
    ext = Path(filename or "").suffix.lower()
    if ext not in ROSTER_EXTENSIONS:
        raise RosterFileError(f"CSV 또는 엑셀(.xlsx) 파일만 올릴 수 있습니다. (받은 파일: {filename})")
    if len(content) > ROSTER_MAX_BYTES:
        raise RosterFileError("파일이 너무 큽니다. 5MB 이하로 올려 주세요.")
    if ext == ".xlsx" and not content.startswith(b"PK"):
        raise RosterFileError("엑셀 파일이 아닙니다. 확장자만 .xlsx 로 바꾼 파일일 수 있습니다.")

    rows = _read_xlsx(content) if ext == ".xlsx" else _read_csv(content)
    if len(rows) < 2:
        raise RosterFileError("첫 줄(열 이름)과 사람 한 명 이상이 있어야 합니다.")

    # 첫 몇 줄 안에 열 이름 줄이 있을 수 있다(제목 줄이 위에 붙은 경우).
    # 아이디 열이 없으면 사번을 아이디로 쓴다(포털을 안 쓰는 명단도 그대로 올라가게).
    cols, start = {}, 0
    for i, row in enumerate(rows[:5]):
        found = _find_columns(row)
        if ("id" in found or "emp_no" in found) and "name" in found:
            cols, start = found, i + 1
            break
    if not cols:
        raise RosterFileError(
            "첫 줄에서 ‘아이디’(또는 ‘사번’)와 ‘이름’ 열을 찾지 못했습니다. "
            "열 이름을 아이디 / 이름 / 부서 / 사번 으로 적어 주세요. (사번·담당자 열은 선택)"
        )

    people, seen, dupes = [], set(), []
    emp_seen, emp_dupes = set(), []
    for row in rows[start:]:
        def cell(key: str) -> str:
            i = cols.get(key)
            return str(row[i]).strip() if i is not None and i < len(row) else ""

        def fix(v: str) -> str:
            # 엑셀이 번호를 숫자로 읽어 '1001.0' 이 되는 일이 잦다. 꼬리를 떼어 준다.
            return v[:-2] if re.fullmatch(r"\d+\.0", v) else v

        uid = fix(cell("id"))
        emp = fix(cell("emp_no"))
        name = cell("name")
        if not uid:
            uid = emp                      # 아이디 열이 없으면 사번을 아이디로
        if not uid or not name:
            continue                       # 빈 줄·합계 줄은 조용히 건너뛴다
        if uid in seen:
            dupes.append(uid)
            continue
        seen.add(uid)
        if emp:
            if emp in emp_seen:
                emp_dupes.append(emp)
            emp_seen.add(emp)
        people.append({
            "id": uid, "emp_no": emp or None, "name": name, "dept": cell("dept"),
            "is_admin": 1 if cell("admin").strip().lower() in TRUE_WORDS else None,
            "active": 1,
        })

    if not people:
        raise RosterFileError("읽을 수 있는 사람이 한 명도 없습니다. 사번과 이름이 채워져 있는지 확인해 주세요.")
    if dupes:
        raise RosterFileError(
            f"아이디가 겹칩니다: {', '.join(dupes[:5])}"
            + (f" 외 {len(dupes) - 5}건" if len(dupes) > 5 else "")
            + ". 아이디는 한 사람에 하나여야 합니다."
        )
    if emp_dupes:
        # 사번으로도 로그인할 수 있게 해 두었으므로, 겹치면 누구인지 가릴 수 없다.
        raise RosterFileError(
            f"사번이 겹칩니다: {', '.join(emp_dupes[:5])}"
            + (f" 외 {len(emp_dupes) - 5}건" if len(emp_dupes) > 5 else "")
            + ". 사번으로도 로그인할 수 있어서 겹치면 안 됩니다."
        )
    return people
