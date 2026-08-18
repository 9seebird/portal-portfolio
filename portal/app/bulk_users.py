"""계정을 한 번에 만든다.

### 왜 필요한가

직원이 500명이면 화면에서 하나씩 만들 수 없다. 인사 담당자가 이미
엑셀로 명부를 들고 있으므로, 그 파일을 그대로 받는 편이 맞다.

### 두 걸음으로 나눈다

    ① 검사  파일을 읽고 **한 줄씩 판정만** 한다. 아무것도 안 만든다.
    ② 적용  검사에서 통과한 줄만 실제로 만든다.

한 번에 하지 않는 이유 — 500줄짜리 파일에 오타가 몇 개 있는 것이
보통이다. 그대로 밀어 넣으면 "몇 명이 들어갔고 누가 빠졌는지"를
나중에 뒤져야 한다. 먼저 보여주고, 보고 나서 누르게 한다.

### 무엇을 받나

    아이디 | 이름 | 부서 | 관리자 | 초기 비밀번호
    ----------------------------------------------
    필수    필수   선택   선택     선택(비우면 자동)

머리글 이름은 여러 가지로 쓸 수 있게 해 두었다. 인사팀 파일이
"사번", "성명", "소속" 이라고 되어 있는 것을 굳이 고치게 할 이유가 없다.

관리자 칸은 O·Y·예·1·TRUE 를 모두 참으로 본다. 사람이 채우는 칸이다.

### 비밀번호

비워 두면 만들어서 돌려준다. 어차피 **첫 로그인에 바꾸게 되므로**
외우기 쉬운 것으로 충분하다. 만든 값은 결과 표에 보여줘서 그대로
전달할 수 있게 한다. (그 뒤로는 다시 볼 수 없다 — 저장하지 않는다)

### 부서

명부에 있는데 포털에 없는 부서는 **말없이 만들지 않는다.**
"영업1팀"과 "영업 1팀"이 따로 생기면 접근 범위가 조용히 갈라진다.

그렇다고 500명짜리 명부를 놓고 부서 스무 개를 손으로 먼저 만들게 하는
것도 말이 안 된다. 그래서 이렇게 한다 — **검사할 때 "없는 부서는 이것들
입니다" 라고 이름을 늘어놓고, 그것까지 만들지 말지 사람이 고르게 한다.**

말없이 만드는 것과 못 만드는 것 사이다. 오타는 눈에 띄고("영업1팀" 과
"영업 1팀" 이 나란히 보이면 바로 안다), 손으로 스무 번 만들 일도 없다.
"""

import io
import re

from . import depts, users

# 머리글로 인정하는 말들. 인사팀 파일을 고치지 않고 그대로 받기 위한 것.
COLUMNS = {
    "user_id": ("아이디", "id", "userid", "user_id", "사번", "계정", "로그인id"),
    "name":    ("이름", "성명", "name", "직원명", "사용자명"),
    "dept":    ("부서", "소속", "부서명", "dept", "department", "팀"),
    "admin":   ("관리자", "admin", "is_admin", "권한", "관리자여부"),
    "password": ("비밀번호", "초기비밀번호", "password", "pw", "암호"),
}

TRUE_WORDS = {"o", "y", "yes", "true", "1", "예", "관리자", "v", "ㅇ"}

# 아이디로 쓸 수 있는 글자. 영문·숫자·. _ - 만.
# 한글 아이디는 로그인 화면에서 입력기 전환 때문에 오히려 불편하다.
ID_RE = re.compile(r"^[A-Za-z0-9._-]{2,32}$")

TEMPLATE_HEADERS = ["아이디", "이름", "부서", "관리자", "초기 비밀번호"]
TEMPLATE_SAMPLE = [
    ["hong", "홍길동", "인사총무팀", "", ""],
    ["kimcs", "김철수", "영업팀", "", "Apple1234!"],
    ["shinys", "신윤성", "인사총무팀", "O", ""],
]


class BulkError(Exception):
    """파일 자체를 못 읽은 경우. 화면에 그대로 보여준다."""


def _norm(s) -> str:
    return re.sub(r"[\s_-]+", "", str(s or "")).strip().lower()


def _rows_from_csv(blob: bytes) -> list[list[str]]:
    import csv
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = blob.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise BulkError("파일 인코딩을 알 수 없습니다. CSV 를 UTF-8 로 저장해 주세요.")
    return [[(c or "").strip() for c in r] for r in csv.reader(io.StringIO(text))]


def _rows_from_xlsx(blob: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise BulkError("엑셀을 읽는 기능이 설치되어 있지 않습니다. (openpyxl)")
    try:
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        raise BulkError(f"엑셀을 열지 못했습니다. .xlsx 로 저장했는지 확인해 주세요. ({e})")
    ws = wb.worksheets[0]
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append([("" if v is None else str(v)).strip() for v in row])
    wb.close()
    return out


def _read(filename: str, blob: bytes) -> list[list[str]]:
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    if ext == "csv":
        return _rows_from_csv(blob)
    if ext in ("xlsx", "xlsm"):
        return _rows_from_xlsx(blob)
    raise BulkError(f"'{ext}' 형식은 읽지 못합니다. 엑셀(.xlsx) 이나 CSV 로 올려주세요.")


def _map_headers(header: list[str]) -> dict[str, int]:
    """머리글 줄에서 각 칸이 몇 번째인지 찾는다."""
    found: dict[str, int] = {}
    for i, cell in enumerate(header):
        n = _norm(cell)
        for key, names in COLUMNS.items():
            if key in found:
                continue
            if n in {_norm(x) for x in names}:
                found[key] = i
    return found


def check(filename: str, blob: bytes, create_depts: bool = False) -> dict:
    """파일을 읽고 한 줄씩 판정한다. **아무것도 만들지 않는다.**

    create_depts 는 "없는 부서도 같이 만들 생각이다" 라는 뜻이다.
    켜면 그 줄들을 문제로 세지 않는다. 실제로 만드는 것은 apply 다.
    """
    rows = _read(filename, blob)
    rows = [r for r in rows if any((c or "").strip() for c in r)]
    if not rows:
        raise BulkError("내용이 없는 파일입니다.")

    cols = _map_headers(rows[0])
    if "user_id" not in cols or "name" not in cols:
        raise BulkError(
            "머리글에서 '아이디'와 '이름' 칸을 찾지 못했습니다. "
            "양식을 내려받아 그 머리글을 그대로 쓰시면 됩니다."
        )

    known_depts = set(depts.all_depts())
    seen: set[str] = set()
    out = []

    for n, row in enumerate(rows[1:], start=2):      # 2 = 엑셀에서 보이는 줄 번호
        def cell(key: str) -> str:
            i = cols.get(key)
            return (row[i].strip() if i is not None and i < len(row) else "")

        uid, name = cell("user_id"), cell("name")
        dept, pw = cell("dept"), cell("password")
        admin = _norm(cell("admin")) in TRUE_WORDS

        item = {"line": n, "user_id": uid, "name": name, "dept": dept,
                "is_admin": admin, "password": pw, "problem": ""}

        if not uid or not name:
            item["problem"] = "아이디와 이름은 비울 수 없습니다"
        elif not ID_RE.match(uid):
            item["problem"] = "아이디는 영문·숫자·. _ - 만, 2~32자"
        elif uid.lower() in seen:
            item["problem"] = "이 파일 안에 같은 아이디가 또 있습니다"
        elif users.get(uid):
            item["problem"] = "이미 있는 계정입니다"
        elif dept and dept not in known_depts and not create_depts:
            # 그냥 만들어 버리지 않는다. 오타가 그대로 부서가 되면
            # 접근 범위가 조용히 갈라진다. 화면에서 켜야 넘어간다.
            item["problem"] = "등록되지 않은 부서입니다 — 아래에서 함께 만들거나, 부서 관리에서 먼저 만드세요"
        elif pw and users.check_password(pw):
            item["problem"] = users.check_password(pw)

        # 새로 생길 부서라고 표시해 둔다. 화면에서 눈에 띄게 그린다.
        if dept and dept not in known_depts:
            item["new_dept"] = True

        if not item["problem"]:
            seen.add(uid.lower())
            # 비어 있으면 만들어 준다. 결과 표에 보여줘서 전달하게 한다.
            item["password"] = pw or users.suggest_password()
        out.append(item)

    ok = [r for r in out if not r["problem"]]

    # 명부에 나온 순서대로. 가나다순으로 흩어 놓으면 오타 짝("영업1팀"/"영업 1팀")이
    # 서로 떨어져서 눈에 안 띈다.
    new_depts: list[str] = []
    for r in out:
        d = r.get("dept")
        if d and d not in known_depts and d not in new_depts:
            new_depts.append(d)

    return {
        "rows": out,
        "ok": len(ok),
        "bad": len(out) - len(ok),
        "admins": sum(1 for r in ok if r["is_admin"]),
        "new_depts": new_depts,
    }


def apply(rows: list[dict], create_depts: bool = False) -> dict:
    """검사에서 통과한 줄만 실제로 만든다.

    화면이 보낸 값을 그대로 믿지 않는다. 만들기 직전에 한 번 더 본다.
    검사와 적용 사이에 다른 관리자가 같은 아이디를 만들었을 수도 있다.

    create_depts 가 켜져 있으면 **계정보다 부서를 먼저** 만든다.
    순서가 뒤바뀌면 첫 사람이 "부서 없음"으로 걸린다.
    """
    made_depts: list[str] = []
    if create_depts:
        have = set(depts.all_depts())
        for r in rows or []:
            d = (r.get("dept") or "").strip()
            if d and d not in have:
                if depts.create(d):
                    made_depts.append(d)
                have.add(d)

    known_depts = set(depts.all_depts())
    made, failed = [], []
    for r in rows or []:
        uid = (r.get("user_id") or "").strip()
        name = (r.get("name") or "").strip()
        dept = (r.get("dept") or "").strip()
        pw = (r.get("password") or "").strip()
        why = ""
        if not uid or not name or not ID_RE.match(uid):
            why = "아이디·이름이 올바르지 않습니다"
        elif users.get(uid):
            why = "이미 있는 계정입니다"
        elif dept and dept not in known_depts:
            why = "등록되지 않은 부서입니다"
        elif not pw or users.check_password(pw):
            why = "초기 비밀번호가 규칙에 맞지 않습니다"

        if why:
            failed.append({**r, "problem": why})
            continue
        try:
            users.create(uid, name, pw, dept=dept,
                         is_admin=bool(r.get("is_admin")), must_change=True)
            made.append({"user_id": uid, "name": name, "dept": dept,
                         "is_admin": bool(r.get("is_admin")), "password": pw})
        except Exception as e:  # noqa: BLE001
            failed.append({**r, "problem": f"만들지 못했습니다 ({type(e).__name__})"})
    return {"made": made, "failed": failed, "made_depts": made_depts}


def template_csv() -> bytes:
    """양식 파일. 엑셀에서 열리도록 BOM 을 붙인다.

    BOM 이 없으면 엑셀이 CP949 로 읽어서 머리글이 깨진다.
    """
    lines = [",".join(TEMPLATE_HEADERS)]
    lines += [",".join(r) for r in TEMPLATE_SAMPLE]
    body = "\n".join(lines) + "\n"
    return "﻿".encode("utf-8") + body.encode("utf-8")
